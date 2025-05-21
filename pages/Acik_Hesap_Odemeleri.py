import streamlit as st
import pandas as pd
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from openpyxl import load_workbook
from DatabaseConnector.mysqlConnector import get_all_acik_hesap_odemeleri

st.set_page_config(page_title="Açık Hesap Ödeme Geçmişi", page_icon="📜", layout="wide")
st.title("📜 Açık Hesap Ödeme Geçmişi")

# Veritabanından ödemeleri çek
odemeler = get_all_acik_hesap_odemeleri()

if not odemeler:
    st.info("🔎 Henüz ödeme kaydı yok.")
else:
    # DataFrame oluştur
    df = pd.DataFrame(odemeler)

    # Beklenen kolonlar:
    # 'customer_name', 'hesap_id', 'customer_number', 'payment', 'payment_type', 'created_at', 'kur'

    # Ödeme Tutarı kur bilgisiyle çarpılarak gösterilir
    df["Ödeme Tutarı (₺)"] = df["payment"] 

    # Kolon adlarını düzenle
    df.rename(columns={
        "customer_name": "Müşteri Adı",
        "hesap_id": "Hesap ID",
        "customer_number": "Telefon",
        "payment_type": "Ödeme Türü",
        "created_at": "Tarih",
        "kur": "Kur",
    }, inplace=True)

    # Sütun sıralaması
    df = df[["Müşteri Adı", "Telefon", "Hesap ID", "Kur", "Ödeme Tutarı (₺)", "Ödeme Türü", "Tarih"]]

    # Filtreleme
    customer_filter = st.text_input("Müşteri Adı veya Telefon ile ara:")
    if customer_filter:
        df = df[
            df["Müşteri Adı"].str.contains(customer_filter, case=False, na=False) |
            df["Telefon"].str.contains(customer_filter, case=False, na=False)
        ]

    # Tarihi datetime'a çevir
    df["Tarih"] = pd.to_datetime(df["Tarih"], errors='coerce')

    # Görüntüleme
    st.dataframe(df, use_container_width=True)

    # Excel çıktısı için hazırlık
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Ödemeler')

    # Excel dosyasını stil vermek için tekrar aç
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb.active

    # Tarih formatı oluştur
    date_style = NamedStyle(name="datetime", number_format='YYYY-MM-DD HH:MM:SS')

    tarih_col_idx = df.columns.get_loc("Tarih") + 1

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=tarih_col_idx)
        cell.style = date_style

    # Son halini kaydet
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    # Download butonu
    st.download_button(
        label="📩 Excel Olarak İndir",
        data=output_buffer,
        file_name="odeme_gecmisi.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
